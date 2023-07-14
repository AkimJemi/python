from openpyxl import load_workbook
import Util

dbFileName = "db/db.xlsx"
pyName = "initialize_LOLTeam_DB"
wb = load_workbook(dbFileName)
tmp = "tmp"
sheet_names = wb.sheetnames
try:
    tmp_sheet = wb[tmp]
except:
    wb.create_sheet(tmp, len(sheet_names))

wb.save(dbFileName)
for sheet_name in sheet_names:
    if sheet_name == tmp:
        continue
    idx = sheet_names.index(sheet_name)
    wb.remove(wb[sheet_name])

sheet_names = Util.sheetnames
for sheet_name in sheet_names:
    idx = sheet_names.index(sheet_name)
    wb.create_sheet(sheet_name, idx)

wb.remove(wb[tmp])
wb.save(dbFileName)
