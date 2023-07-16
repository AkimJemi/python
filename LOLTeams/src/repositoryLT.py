from openpyxl import load_workbook
import Util as u
u.showLog(__file__, None, None)
dbFileName = u.dbFileName


def sheetAutoCheckAndInsert(info, sheetName, num):
    for i in info:
        if i == '' or i == None:
            ms

    u.showLog(__file__, sheetAutoCheckAndInsert.__name__,
              "info : {}, sheetName : {}, num : {}".format(info, sheetName, num))
    wb = load_workbook(dbFileName)
    ws = wb.active
    ws_names = wb.sheetnames
    sheet = None
    for a_sheet in ws_names:
        if a_sheet == sheetName:
            sheet = wb.get_sheet_by_name(sheetName)

    if sheet == None:
        wb.create_sheet(sheetName)
        sheet = wb[sheetName]

    max_row = sheet.max_row
    if sheet['A1'].value == None and max_row == 1:
        max_row = 1
    elif sheet['A1'].value != None and max_row == 1:
        max_row = 2
    else:
        max_row = max_row + 1

    sheet.cell(row=max_row, column=1).value = max_row
    for i in range(2, num+2):
        sheet.cell(row=max_row, column=i).value = info[i-2]
    wb.save(dbFileName)


def getUserList():
    u.showLog(__file__, getUserList.__name__, None)
    sheetName = u.sheetnames[1]
    wb = load_workbook(filename=dbFileName, data_only=True)
    ws = wb[sheetName]
    userList = []
    for y in range(1, ws.max_row+1):
        userList.append(ws.cell(column=2, row=y).value)

    return userList


def saveGameMatchTeamSetting(info):
    u.showLog(__file__, saveGameMatchTeamSetting.__name__, info)
    sheetName = u.sheetnames[0]
    sheetAutoCheckAndInsert(info, sheetName, 5)


def saveUserInfo(info):
    u.showLog(__file__, saveUserInfo.__name__, info)
    sheetName = u.sheetnames[1]
    sheetAutoCheckAndInsert(info, sheetName, 3)


def saveUserInfo(info):
    u.showLog(__file__, saveUserInfo.__name__, info)
    sheetName = u.sheetnames[1]
    sheetAutoCheckAndInsert(info, sheetName, 3)
    from mainPage import setUserList
